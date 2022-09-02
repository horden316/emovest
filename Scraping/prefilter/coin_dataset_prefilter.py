import openpyxl
import os

#讀入要處理的資料集
basepath=os.getcwd()
workbook_location = basepath + '/Scraping/prefilter/0818評論-空白.xlsx'
workbook = openpyxl.load_workbook(workbook_location)
sheet = workbook.worksheets[0]

#讀入黑名單config
blacklist_config_location = basepath + "/Scraping/prefilter/blacklist.config"
with open(blacklist_config_location,"r") as f:
    blacklist = f.read().splitlines()

#讀入第一個字黑名單config
firstword_blacklist_config_location = basepath + "/Scraping/prefilter/firstword_blacklist.config"
with open(firstword_blacklist_config_location,"r") as f:
    firstword_blacklist = f.read().splitlines()

#讀入$後面接的幣種白名單config
coin_whitelist_config_location=basepath + "/Scraping/prefilter/coin_whitelist.config"
with open(coin_whitelist_config_location,"r") as f:
    coin_whitelist = f.read().splitlines()

deleted_row_count=0

#print(sheet['C3'].value)

#print(len(blacklist))
#print(blacklist[1])

#print(sheet.max_row)


#Simple Blacklist
for k in range(len(blacklist)):
    for i in range(2,sheet.max_row):
        #print("k="+str(k)+"  i="+str(i))
        if blacklist[k] in str(sheet['C'+str(i)].value):
            sheet.delete_rows(i, 1)
            deleted_row_count+=1
            print("delete row: "+str(i)+" because of "+str(blacklist[k]))

workbook.save(workbook_location.strip(".xlsx")+"_clean"+".xlsx")

#5W1H
for k in range(len(firstword_blacklist)):
    for i in range(2,sheet.max_row):
        temp=(str(sheet['C'+str(i)].value)).split()
        if (firstword_blacklist[k]==temp[0]):
            sheet.delete_rows(i, 1)
            deleted_row_count+=1
            print("delete row: "+str(i)+" because of "+str(firstword_blacklist[k]))

workbook.save(workbook_location.strip(".xlsx")+"_clean"+".xlsx")



#$ with only whitelist
for i in range(2,sheet.max_row):
    temp=(str(sheet['C'+str(i)].value)).split()
    trash=False
    for j in range(len(temp)):
        if(((temp[j])[0]=="$") and (((temp[j])[1:])not in coin_whitelist)):
            trash=True
        if(trash==True):
            print("delete row: "+str(i)+" because of illegal $coin "+str((temp[j])[1:])+"|")
            sheet.delete_rows(i, 1)
            deleted_row_count+=1
            break

        

workbook.save(workbook_location.strip(".xlsx")+"_clean"+".xlsx")

#remove non english

#numbers + today/weekend


print("total deleted: "+ str(deleted_row_count))